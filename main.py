import tkinter as tk
from tkinter import simpledialog, messagebox, filedialog
from datetime import datetime
import mysql.connector
from openpyxl import Workbook
import unicodedata
import logging
from tkinter import ttk

# Função para remover acentos
def remover_acentos(texto):
    return ''.join(ch for ch in unicodedata.normalize('NFD', texto) if unicodedata.category(ch) != 'Mn')

# Conectar ao banco de dados
connection = mysql.connector.connect(
    host="localhost",
    user="root",
    password="bacana",
    database="relogio_ponto"
)

# Função para buscar funcionários no banco de dados
def buscar_funcionarios():
    cursor = connection.cursor()
    cursor.execute("SELECT codigo, nome FROM Funcionarios")
    funcionarios = {codigo: nome for codigo, nome in cursor}
    cursor.close()
    return funcionarios

# Função para determinar o turno com base no horário
def determinar_turno(hora):
    # Separar as horas e os minutos do horário
    horas, minutos = map(int, hora.split(':'))

    # Calcular o total de minutos desde a meia-noite
    total_minutos = horas * 60 + minutos

    # Definir os limites dos turnos em minutos
    limite_manha_inicio = 5 * 60
    limite_manha_fim = 12 * 60 + 30  # Fim da manhã: 12:30 (meio-dia)
    limite_tarde_inicio = 12 * 60 + 30  # Início da tarde: 12:30 (meio-dia)
    limite_tarde_fim = 18 * 60  # Fim da tarde: 18:00 (6 horas da tarde)
    limite_noite_inicio = 18 * 60  # Início da noite: 18:00 (6 horas da tarde)
    limite_noite_fim = 24 * 60 + 5 * 60  # Fim da noite: 5:00 (5 horas da manhã do outro dia)

    # Determinar o turno com base no total de minutos
    if limite_manha_inicio <= total_minutos < limite_manha_fim:
        return 'Manhã'
    elif limite_tarde_inicio <= total_minutos < limite_tarde_fim:
        return 'Tarde'
    elif limite_noite_inicio <= total_minutos < limite_noite_fim:
        return 'Noite'
    else:
        return 'Extra (Noite)'

# Função para registrar entrada ou saída de turno no banco de dados
def registrar_ponto(event):
    cursor = connection.cursor()
    codigo_funcionario = simpledialog.askinteger("Registrar Ponto", "Digite o código do funcionário:", parent=root)
    if codigo_funcionario:
        cursor.execute("SELECT nome FROM Funcionarios WHERE codigo = %s", (codigo_funcionario,))
        resultado = cursor.fetchone()
        if resultado:
            nome_funcionario = resultado[0]
            hora_registro = datetime.now().strftime("%H:%M")
            logging.info("Hora atual (antes de formatar): %s", datetime.now())
            logging.info("Hora atual (após de formatar): %s", hora_registro)
            turno = determinar_turno(hora_registro)
            tipo_registro = 'Entrada' if obter_ultimo_registro(cursor, codigo_funcionario) != 'Entrada' else 'Saída'
            cursor.execute("INSERT INTO RegistrosPonto (codigo_funcionario, tipo_ponto, data, hora) VALUES (%s, %s, CURDATE(), %s)",
                           (codigo_funcionario, tipo_registro, hora_registro))
            connection.commit()
            
            # Criar uma nova janela temporária para exibir o registro
            popup = tk.Toplevel(root)
            popup.title("Registro")
            popup.geometry("400x100")
            # Exibir registro de horário em um tk.Label na nova janela
            registro_msg = f"{tipo_registro} registrada para {nome_funcionario} no turno da {turno} às {hora_registro}"
            label_registro = tk.Label(popup, text=registro_msg)
            label_registro.pack(pady=20)
            # Destruir a janela pop-up após 1 segundo
            root.after(2000, popup.destroy)
        else:
            messagebox.showerror("Erro", "Código de funcionário inválido.")
    cursor.close()

# Função para obter o último registro de um funcionário
def obter_ultimo_registro(cursor, codigo_funcionario):
    cursor.execute("SELECT tipo_ponto FROM RegistrosPonto WHERE codigo_funcionario = %s ORDER BY id DESC LIMIT 1", (codigo_funcionario,))
    resultado = cursor.fetchone()
    if resultado:
        return resultado[0]
    else:
        return None

# Função para gerar relatório
def gerar_relatorio():
    try:
        cursor = connection.cursor()
        cursor.execute("SELECT f.codigo, f.nome, rp.data, rp.hora, rp.tipo_ponto FROM Funcionarios f INNER JOIN RegistrosPonto rp ON f.codigo = rp.codigo_funcionario ORDER BY f.codigo, rp.data")

        relatorio = cursor.fetchall()

        # Verificar se há dados disponíveis para gerar o relatório
        if not relatorio:
            messagebox.showwarning("Aviso", "Não há dados disponíveis para gerar o relatório.")
            return

        # Solicitar ao usuário o local para salvar o relatório
        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Arquivo Excel", "*.xlsx")])
        if file_path:
            # Criar arquivo Excel
            wb = Workbook()
            for linha in relatorio:
                codigo_funcionario, nome_funcionario, data_str, hora, tipo_ponto = linha
                nome_planilha = remover_acentos(nome_funcionario)
                if nome_planilha not in wb.sheetnames:
                    ws = wb.create_sheet(title=nome_planilha)
                    ws.append(["Data", "Hora", "Turno", "Tipo de Ponto"])  # Adicionar cabeçalhos da planilha
                
                # Converter a string de data em um objeto datetime
                data = datetime.strptime(data_str, "%Y-%m-%d")

                # Formatar a data para "dd/mm/yyyy"
                data_formatada = data.strftime("%d/%m/%Y")
                
                # Determinar o turno com base na hora
                turno = determinar_turno(hora)
                
                # Adicionar entrada ou saída ao tipo de ponto
                tipo_ponto_label = "Entrada" if tipo_ponto == "Entrada" else "Saída"
                
                ws.append([data_formatada, hora, turno, tipo_ponto_label])  # Adicionar dados à planilha

                # Verificar se os dados de data/hora estão sendo recuperados corretamente
                print("Data/Hora:", data, hora)

            try:
                wb.remove(wb["Sheet"])  # Remover a planilha padrão criada automaticamente
                wb.save(file_path)
            except Exception as e:
                messagebox.showerror("Erro", f"Erro ao salvar o arquivo Excel: {e}")

            # Limpar apenas os dados de registros do banco de dados
            cursor.execute("DELETE FROM RegistrosPonto")
            connection.commit()

            messagebox.showinfo("Relatório", "Relatório gerado com sucesso.")
    except Exception as e:
        messagebox.showerror("Erro", f"Erro ao gerar relatório: {e}")
    finally:
        if cursor:
            cursor.close()

# Função para limpar os registros
def limpar_registros():
    cursor = connection.cursor()
    cursor.execute("DELETE FROM RegistrosPonto")
    connection.commit()
    cursor.close()

# Função para atualizar o relógio
def atualizar_relogio(label_relogio):
    hora_atual = datetime.now().strftime("%H:%M:%S")
    data_atual = datetime.now().strftime("%d/%m/%Y")
    label_relogio.config(text=f"Hora: {hora_atual}\nData: {data_atual}")
    root.after(1000, atualizar_relogio, label_relogio)

# Função principal para criar a interface gráfica
def criar_interface():
    global root
    root = tk.Tk()
    root.title("Relógio Ponto")

    # Define a janela para ser maximizada
    root.state('zoomed')

    # Label para exibir o relógio com a data
    label_relogio = tk.Label(root, font=("Arial", 60, 'bold'))
    label_relogio.grid(row=0, column=0, columnspan=2, padx=20, pady=20, sticky="nsew")
    atualizar_relogio(label_relogio)

    # Estilo personalizado para o botão com tamanho aumentado
    style = ttk.Style()
    style.configure("TButton", padding=(25, 15))  # Ajuste os valores conforme necessário

    # Frame para os botões
    frame_botoes = tk.Frame(root)
    frame_botoes.grid(row=1, column=0, padx=20, pady=20, sticky="nw")

    # Criar o estilo para o botão com a fonte desejada
    style = ttk.Style()
    style.configure("TButton", font=("Arial", 15, ))
    
    # Botão para gerar relatório com estilo personalizado
    btn_relatorio = ttk.Button(frame_botoes, text="Gerar Relatório", command=gerar_relatorio, style='TButton')
    btn_relatorio.grid(row=0, column=0, padx=20, pady=20)

    # Frame para a mensagem
    frame_mensagem = tk.Frame(root)
    frame_mensagem.grid(row=2, column=0, columnspan=2, padx=20, pady=20)

    # Label para exibir a mensagem de registro de horário
    label_mensagem = tk.Label(frame_mensagem, text="Pressione 1 para registrar horário", font=("Arial", 40))
    label_mensagem.pack()

    # Configurar evento de tecla para chamar a função registrar_ponto quando a tecla "1" for pressionada
    root.bind("<Key>", registrar_ponto)

    # Centralizando horizontalmente e verticalmente o relógio
    root.grid_rowconfigure(0, weight=1)
    root.grid_columnconfigure(0, weight=1)

    # Centralizando a mensagem na parte inferior
    frame_mensagem.grid_rowconfigure(0, weight=1)
    frame_mensagem.grid_columnconfigure(0, weight=1)

    # Focar na janela principal para garantir que o evento de tecla seja capturado corretamente
    root.focus_set()

    root.mainloop()

# Chamando a função principal para criar a interface gráfica
criar_interface()
